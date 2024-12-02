import openpyxl
import time

# Hàm lấy thời gian hiện tại
def getTimeNow():
    utc_time = time.gmtime()
    utc_plus_7_time = time.localtime(time.mktime(utc_time) + 7 * 3600)
    time_now = time.strftime('%d-%m-20%y | %H:%M:%S', utc_plus_7_time)
    return time_now

# Hàm đọc dữ liệu từ file Excel
def getData(file_path="data.xlsx"):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Kiểm tra nếu cột 'id' không rỗng
            data.append(list(row))
    wb.close()
    return data

# Hàm thêm dữ liệu mới vào file Excel
def postData(info, file_path="data.xlsx"):
    mst, name, note = info
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Tìm số dòng dữ liệu thực tế, bỏ qua tiêu đề (dòng 1)
    last_row = sheet.max_row
    if last_row < 2 or not sheet.cell(row=2, column=1).value:  # Nếu chỉ có tiêu đề hoặc hàng 2 rỗng
        new_id = 1
    else:
        new_id = last_row  # ID mới là dòng cuối chứa dữ liệu + 1

    new_data = [new_id, mst, name, note, getTimeNow()]
    sheet.append(new_data)
    wb.save(file_path)
    wb.close()
    return {"status": "success", "message": "Data added successfully", "id": new_id}

# Hàm chỉnh sửa dữ liệu dựa trên MST
def editData(info, file_path="data.xlsx"):
    mst, name, note = info
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    found = False
    for row in sheet.iter_rows(min_row=2):
        if str(row[1].value) == str(mst):
            row[2].value = name
            row[3].value = note
            row[4].value = getTimeNow()
            found = True
            break

    if found:
        wb.save(file_path)
        result = {"status": "success", "message": "Data updated successfully", "mst": mst}
    else:
        result = {"status": "error", "message": "MST not found"}

    wb.close()
    return result

# print(getData())
# data = ('12 34', 'sdfgh7', '12456789')
# print(editData(data))
# print(postData(data))